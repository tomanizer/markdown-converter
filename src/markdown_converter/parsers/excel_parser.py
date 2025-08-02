"""
Excel Document Parser

This module provides specialized parsing for Excel documents (.xlsx, .xls)
using openpyxl for spreadsheet processing and pandas for data manipulation.
"""

import logging
import os
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from ..core.exceptions import ParserError, UnsupportedFormatError
from .base import BaseParser, ParserResult


class ExcelParser(BaseParser):
    """
    Specialized parser for Excel documents (.xlsx, .xls).

    Uses openpyxl for spreadsheet processing and pandas for data manipulation.
    """

    def __init__(self, config: Optional[Dict[str, Any]] = None) -> None:
        """
        Initialize the Excel parser.

        :param config: Configuration dictionary
        """
        super().__init__(config)
        self.logger = logging.getLogger("ExcelParser")
        self._setup_default_config()
        self._validate_dependencies()

    def _setup_default_config(self) -> None:
        """Setup default configuration for Excel parsing."""
        self.default_config = {
            # openpyxl options
            "openpyxl_options": {
                "data_only": True,  # Read values, not formulas
                "keep_vba": False,
                "read_only": True,
            },
            # pandas options
            "pandas_options": {
                "header": 0,  # Use first row as header
                "index_col": None,
                "na_values": ["", "NA", "N/A"],
                "keep_default_na": True,
            },
            # Output settings
            "include_all_sheets": True,
            "max_sheets": 10,  # Limit number of sheets to process
            "convert_tables": True,
            "include_formulas": False,
            "include_charts": False,
            "preserve_formatting": False,
        }

        # Merge with user config
        if self.config:
            self.default_config.update(self.config)

    def _validate_dependencies(self) -> None:
        """Validate that required dependencies are available."""
        try:
            import openpyxl

            self.openpyxl_available = True
        except ImportError:
            self.openpyxl_available = False
            self.logger.warning("openpyxl not available")

        try:
            import pandas

            self.pandas_available = True
        except ImportError:
            self.pandas_available = False
            self.logger.warning("pandas not available")

        if not self.openpyxl_available and not self.pandas_available:
            raise ParserError("Neither openpyxl nor pandas are available")

    def can_parse(self, file_path: Union[str, Path]) -> bool:
        """
        Check if this parser can handle the given file.

        :param file_path: Path to the file
        :return: True if the file can be parsed
        """
        file_path = Path(file_path)
        return file_path.suffix.lower() in [".xlsx", ".xls", ".xlsb"]

    def parse(self, file_path: Union[str, Path]) -> ParserResult:
        """
        Parse an Excel document and extract its content.

        :param file_path: Path to the Excel document
        :return: ParserResult with extracted content and metadata
        :raises: ParserError if parsing fails
        """
        file_path = Path(file_path)

        if not self.can_parse(file_path):
            raise UnsupportedFormatError(f"Cannot parse {file_path.suffix} files")

        self.logger.info(f"Parsing Excel document: {file_path}")

        # Check if file needs conversion (e.g., .xlsb -> .xlsx)
        converted_file = self._ensure_readable_format(file_path)

        try:
            # Try pandas first (preferred for data analysis)
            if self.pandas_available:
                try:
                    return self._parse_with_pandas(converted_file)
                except Exception as e:
                    self.logger.warning(f"pandas parsing failed: {e}, trying openpyxl")

            # Fallback to openpyxl
            if self.openpyxl_available:
                return self._parse_with_openpyxl(converted_file)

            raise ParserError("No available Excel parser found")

        except Exception as e:
            self.logger.error(f"Excel parsing failed for {file_path}: {e}")
            raise ParserError(f"Failed to parse Excel document {file_path}: {e}")
        finally:
            # Clean up converted file if it's different from original
            if converted_file != file_path and self.default_config.get(
                "cleanup_temp_files", True
            ):
                self._cleanup_converted_file(converted_file)

    def _parse_with_pandas(self, file_path: Path) -> ParserResult:
        """
        Parse Excel document using pandas.

        :param file_path: Path to the Excel document
        :return: ParserResult with extracted content
        """
        import pandas as pd

        self.logger.debug(f"Using pandas to parse {file_path}")

        content_parts = []
        sheets_data = []
        metadata = {
            "parser": "pandas",
            "file_path": str(file_path),
            "file_size": file_path.stat().st_size,
            "format": file_path.suffix.lower(),
        }

        # Read all sheets
        excel_file = pd.ExcelFile(file_path)
        sheet_names = excel_file.sheet_names

        metadata["sheet_count"] = len(sheet_names)
        metadata["sheet_names"] = sheet_names

        # Process each sheet
        for sheet_name in sheet_names[: self.default_config["max_sheets"]]:
            self.logger.debug(f"Processing sheet: {sheet_name}")

            try:
                # Read sheet data
                df = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    **self.default_config["pandas_options"],
                )

                if not df.empty:
                    # Add sheet header
                    content_parts.append(f"# Sheet: {sheet_name}")
                    content_parts.append("")

                    # Convert DataFrame to markdown table
                    table_markdown = self._dataframe_to_markdown(df)
                    content_parts.append(table_markdown)
                    content_parts.append("")  # Add blank line

                    # Store sheet data
                    sheets_data.append(
                        {
                            "name": sheet_name,
                            "rows": len(df),
                            "columns": len(df.columns),
                            "content": table_markdown,
                        }
                    )

            except Exception as e:
                self.logger.warning(f"Failed to process sheet {sheet_name}: {e}")
                content_parts.append(f"# Sheet: {sheet_name}")
                content_parts.append("*Error processing this sheet*")
                content_parts.append("")

        # Combine all content
        content = "\n".join(content_parts)

        # Add sheets metadata
        if sheets_data:
            metadata["sheets"] = sheets_data

        return ParserResult(
            content=content, metadata=metadata, format="markdown", messages=[]
        )

    def _parse_with_openpyxl(self, file_path: Path) -> ParserResult:
        """
        Parse Excel document using openpyxl.

        :param file_path: Path to the Excel document
        :return: ParserResult with extracted content
        """
        from openpyxl import load_workbook

        self.logger.debug(f"Using openpyxl to parse {file_path}")

        content_parts = []
        sheets_data = []
        metadata = {
            "parser": "openpyxl",
            "file_path": str(file_path),
            "file_size": file_path.stat().st_size,
            "format": file_path.suffix.lower(),
        }

        # Load workbook
        workbook = load_workbook(
            file_path,
            data_only=self.default_config["openpyxl_options"]["data_only"],
            read_only=self.default_config["openpyxl_options"]["read_only"],
        )

        sheet_names = workbook.sheetnames
        metadata["sheet_count"] = len(sheet_names)
        metadata["sheet_names"] = sheet_names

        # Process each sheet
        for sheet_name in sheet_names[: self.default_config["max_sheets"]]:
            self.logger.debug(f"Processing sheet: {sheet_name}")

            try:
                worksheet = workbook[sheet_name]

                # Extract data from worksheet
                sheet_data = self._extract_worksheet_data(worksheet)

                if sheet_data:
                    # Add sheet header
                    content_parts.append(f"# Sheet: {sheet_name}")
                    content_parts.append("")

                    # Convert to markdown table
                    table_markdown = self._convert_data_to_markdown(sheet_data)
                    content_parts.append(table_markdown)
                    content_parts.append("")  # Add blank line

                    # Store sheet data
                    sheets_data.append(
                        {
                            "name": sheet_name,
                            "rows": len(sheet_data),
                            "columns": len(sheet_data[0]) if sheet_data else 0,
                            "content": table_markdown,
                        }
                    )

            except Exception as e:
                self.logger.warning(f"Failed to process sheet {sheet_name}: {e}")
                content_parts.append(f"# Sheet: {sheet_name}")
                content_parts.append("*Error processing this sheet*")
                content_parts.append("")

        workbook.close()

        # Combine all content
        content = "\n".join(content_parts)

        # Add sheets metadata
        if sheets_data:
            metadata["sheets"] = sheets_data

        return ParserResult(
            content=content, metadata=metadata, format="markdown", messages=[]
        )

    def _extract_worksheet_data(self, worksheet) -> List[List[str]]:
        """
        Extract data from a worksheet.

        :param worksheet: openpyxl worksheet object
        :return: List of rows (each row is a list of cell values)
        """
        data = []

        # Get the used range
        for row in worksheet.iter_rows(values_only=True):
            # Convert all values to strings and clean them
            row_data = []
            for cell_value in row:
                if cell_value is None:
                    cell_text = ""
                else:
                    cell_text = str(cell_value).strip()
                row_data.append(cell_text)

            # Only add non-empty rows
            if any(cell.strip() for cell in row_data):
                data.append(row_data)

        return data

    def _dataframe_to_markdown(self, df) -> str:
        """
        Convert a pandas DataFrame to markdown table.

        :param df: pandas DataFrame
        :return: Markdown table string
        """
        if df.empty:
            return "*Empty sheet*"

        # Convert DataFrame to markdown
        markdown_table = df.to_markdown(index=False)

        return markdown_table

    def _convert_data_to_markdown(self, data: List[List[str]]) -> str:
        """
        Convert data to markdown table format.

        :param data: List of rows (each row is a list of cell values)
        :return: Markdown table string
        """
        if not data or not data[0]:
            return "*Empty sheet*"

        markdown_lines = []

        # Process each row
        for i, row in enumerate(data):
            # Clean and escape cell content
            cells = []
            for cell in row:
                if cell is None:
                    cell_text = ""
                else:
                    cell_text = str(cell).strip().replace("|", "\\|")
                cells.append(cell_text)

            # Create markdown row
            markdown_row = "| " + " | ".join(cells) + " |"
            markdown_lines.append(markdown_row)

            # Add separator row after header
            if i == 0:
                separator = "| " + " | ".join(["---"] * len(cells)) + " |"
                markdown_lines.append(separator)

        return "\n".join(markdown_lines)

    def get_supported_formats(self) -> List[str]:
        """
        Get list of supported file formats.

        :return: List of supported format extensions
        """
        return [".xlsx", ".xls", ".xlsb"]

    def _ensure_readable_format(self, file_path: Path) -> Path:
        """
        Ensure the file is in a readable format, converting if necessary.

        :param file_path: Path to the file
        :return: Path to the readable file (original or converted)
        """
        # Check if file needs conversion
        if file_path.suffix.lower() in [".xlsb", ".xls"]:
            try:
                from ..core.file_converter import FileConverter

                converter = FileConverter()
                if converter.needs_conversion(file_path):
                    self.logger.info(f"Converting {file_path} to readable format")
                    return converter.convert_file(file_path)
            except ImportError:
                self.logger.warning(
                    "FileConverter not available, trying direct parsing"
                )
            except Exception as e:
                self.logger.warning(
                    f"File conversion failed: {e}, trying direct parsing"
                )

        return file_path

    def _cleanup_converted_file(self, file_path: Path) -> None:
        """
        Clean up a converted temporary file.

        :param file_path: Path to the file to clean up
        """
        try:
            if file_path.exists():
                file_path.unlink()
                self.logger.debug(f"Cleaned up temporary file: {file_path}")
        except Exception as e:
            self.logger.warning(f"Failed to clean up temporary file {file_path}: {e}")

    def get_parser_info(self) -> Dict[str, Any]:
        """
        Get information about this parser.

        :return: Dictionary with parser information
        """
        return {
            "name": "ExcelParser",
            "description": "Specialized parser for Excel documents",
            "supported_formats": self.get_supported_formats(),
            "dependencies": {
                "openpyxl": self.openpyxl_available,
                "pandas": self.pandas_available,
            },
            "config": self.default_config,
        }
